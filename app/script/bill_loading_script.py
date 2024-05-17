from config import CONFIGURATION as conf

import os
import json
from openpyxl import load_workbook
import shutil

# database_dir = "P:\\app\\database"
working_dir = conf["working_dir"]
database_dir = conf["database_dir"]
accounting_dir = conf["accounting_dir"]
bill_dir = "S:\\20240402-bill_classification\\Project related\\Bill"
mp_dir = os.path.join(database_dir, "mp.json")
mp_json = json.load(open(mp_dir))
count=0


quotation_bill_number_map = {}
wb = load_workbook("S:\\20240402-bill_classification\\bill_classification_result.xlsx")
ws = wb["Sheet1"]

i = 2
while True:
    if ws[f"A{str(i)}"].value is None:
        break

    # print(ws[f"A{str(i)}"].value)
    quotation = ws[f"B{str(i)}"].value
    print(quotation)
    # if quotation is None:
    #     i+=1
    #     continue
    # assert os.path.exists(os.path.join(bill_dir, ws[f"A{str(i)}"].value))
    project_number = mp_json[quotation]["Proj. No."]
    # ws[f"C{str(i)}"].value = project_number
    # if mp_json[quotation]["Proj. No."] is None:
    #     project_address = "P:\\"+quotation+"-"+mp_json[quotation]["Project detailed address"]
    # else:
    #     project_address = "P:\\"+mp_json[quotation]["Proj. No."]+"-"+mp_json[quotation]["Project detailed address"]
    # ws[f"D{str(i)}"].value = project_address
    # ws[f"E{str(i)}"].value = mp_json[quotation]["Service Been Engaged"]
    # assert ws[f"F{str(i)}"].value == "Variation" or ws[f"F{str(i)}"].value in ws[f"E{str(i)}"].value
    # ws[f"L{str(i)}"].value = str(round(ws[f"J{str(i)}"].value*1.1, 1)) if ws[f"K{str(i)}"].value else ws[f"J{str(i)}"].value
    # if not quotation in quotation_bill_number_map.keys():
    #     quotation_bill_number_map[quotation] = 1
    # else:
    #     quotation_bill_number_map[quotation]+=1
    # letter = chr(64+quotation_bill_number_map[quotation])
    letter = ws[f"M{str(i)}"].value

    data_dir = os.path.join(database_dir, quotation, "data.json")
    data_json = json.load(open(data_dir))

    service = ws[f"F{str(i)}"].value
    # if service!="Variation":
    #     assert data_json["Project Info"]["Project"]["Service Type"][service]["Include"]
    from_contact = ws[f"G{str(i)}"].value
    description = ws[f"H{str(i)}"].value + "-"+ws[f"A{str(i)}"].value.rsplit(".", 1)[0]
    type = ws[f"I{str(i)}"].value
    fee = ws[f"J{str(i)}"].value
    no_gst =  not ws[f"K{str(i)}"].value
    in_gst = ws[f"L{str(i)}"].value
    # print(data_json["Bills"]["Details"][service]["Service"])
    # assert data_json["Bills"]["Details"][service]["Service"]=="" or data_json["Bills"]["Details"][service]["Service"]==from_contact

    load = False
    for content in data_json["Bills"]["Details"][service]["Content"]:
        if len(content["Number"])==0:
            assert "Number" in content.keys()
            content["Number"] = letter
            assert "Contact" in content.keys()
            content["Contact"] = from_contact
            content["Service"] = description
            assert "Number" in content.keys()
            content["Fee"] = str(round(fee,1))
            assert "no.GST" in content.keys()
            content["no.GST"] = no_gst
            assert "in.GST" in content.keys()
            content["in.GST"] = str(in_gst)
            assert type in conf["bill_type"]
            assert "Type" in content.keys()
            content["Type"] = type
            content["State"] = "Paid"
            data_json["Bills"]["Details"][service]["no.GST"] = content["no.GST"]
            load = True
            break
    assert load
    data_json["Bills"]["Details"][service]["Service"] = ", ".join(list(set([content["Contact"] for content in data_json["Bills"]["Details"][service]["Content"] if len(content["Contact"])>0])))

    # shutil.copy(os.path.join(bill_dir, ws[f"A{str(i)}"].value),
    #             os.path.join(accounting_dir, quotation, f"{project_number+letter}-"+ws[f"A{str(i)}"].value))
    quote_fee = 0
    quote_ingst = 0
    for content in data_json["Bills"]["Details"][service]["Content"]:
        if len(content["Fee"]) != 0:
            quote_fee += float(content["Fee"])
        if len(content["in.GST"]) != 0:
            quote_ingst += float(content["in.GST"])

    data_json["Bills"]["Details"][service]["Origin"] = str(round(quote_fee, 1))

    data_json["Bills"]["Details"][service]["Fee"] = str(round(quote_fee, 1))
    data_json["Bills"]["Details"][service]["in.GST"] = str(round(quote_ingst, 1))
    data_json["Bills"]["Details"][service]["Paid"] = str(round(quote_fee, 1))
    data_json["Bills"]["Details"][service]["Paid.GST"] = str(round(quote_ingst, 1))

    total_fee = 0
    total_ingst = 0
    for value in data_json["Bills"]["Details"].values():
        if len(value["Fee"]) != 0:
            total_fee += float(value["Fee"])
        if len(value["in.GST"]) != 0:
            total_ingst +=float(value["in.GST"])

    data_json["Bills"]["Fee"] = str(round(total_fee, 1))
    data_json["Bills"]["in.GST"] = str(round(total_ingst,1))

    service_fee = float(data_json["Invoices"]["Details"][service]["Fee"]) if len(data_json["Invoices"]["Details"][service]["Fee"])!=0 else 0
    service_ingst = float(data_json["Invoices"]["Details"][service]["in.GST"]) if len(
        data_json["Invoices"]["Details"][service]["in.GST"]) != 0 else 0

    data_json["Profits"]["Details"][service]["Fee"] = str(round(service_fee-float(data_json["Bills"]["Details"][service]["Fee"]), 1))
    data_json["Profits"]["Details"][service]["in.GST"] = str(round(service_ingst - float(data_json["Bills"]["Details"][service]["in.GST"]), 1))

    data_json["Profits"]["Fee"] = str(round(float(data_json["Invoices"]["Fee"])-float(data_json["Bills"]["Fee"]), 1))
    data_json["Profits"]["in.GST"] = str(round(float(data_json["Invoices"]["in.GST"]) - float(data_json["Bills"]["in.GST"]), 1))


    mp_json[quotation]["Total Bill InGST"] = str(round(total_ingst, 1))

    with open(data_dir, "w") as f:
        json_object = json.dumps(data_json, indent=4)
        f.write(json_object)
    i+=1

with open(mp_dir, "w") as f:
    json_object = json.dumps(mp_json, indent=4)
    f.write(json_object)
i+=1


wb.save('S:\\20240402-bill_classification\\bill_classification_result_xxx.xlsx')




# for dir in os.listdir(database_dir):
#     if os.path.isdir(os.path.join(database_dir, dir)):
#         data_dir = os.path.join(database_dir, dir, "data.json")
#         data_json = json.load(open(data_dir))
