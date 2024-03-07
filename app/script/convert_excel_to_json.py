import json

import openpyxl
from win32com import client as win32client
from openpyxl import load_workbook


from config import CONFIGURATION as conf
from utility import isfloat


import os


def generate_last_inv_to_quotation_map(database_dir):
    res = {}
    for dir in os.listdir(database_dir):
        if os.path.isdir(os.path.join(database_dir, dir)):
            data_json = json.load(open(os.path.join(database_dir, dir, "data.json")))
            inv_list = [inv["Number"] for inv in data_json["Invoices Number"] if len(inv["Number"]) != 0]
            if len(inv_list) != 0:
                res[max(inv_list)] = data_json["Project Info"]["Project"]["Quotation Number"]
    return res

# def convert_asana_name_to_project_number(json_list):
#     return [e.split("\\")[-1].split("-")[0] for e in json_list]

project_name_last_inv_list = []


def check_valid_project(data_json):
    excel_total = 0
    for service in conf["service_list"]:
        if data_json["Invoices"]["Details"][service]["Include"]:
            subtotal = 0
            for content in data_json["Invoices"]["Details"][service]["Content"]:
                subtotal += 0 if len(content["Fee"]) == 0 else float(content["Fee"])
            # assert subtotal !=0
            excel_total += subtotal
            data_json["Invoices"]["Details"][service]["Fee"] = str(subtotal)
            data_json["Invoices"]["Details"][service]["in.GST"] = str(round(subtotal*1.1, 2))
    #varitaion
    subtotal = 0
    for content in data_json["Invoices"]["Details"]["Variation"]["Content"]:
        subtotal += 0 if len(content["Fee"]) == 0 else float(content["Fee"])
    data_json["Invoices"]["Details"]["Variation"]["Fee"] = str(subtotal)
    data_json["Invoices"]["Details"]["Variation"]["in.GST"] = str(round(subtotal * 1.1, 2))
    excel_total += subtotal
    # assert excel_total == float(data_json["Invoices"]["Fee"])


def get_last_inv_number(data_json):
    inv_list = [inv["Number"] for inv in data_json["Invoices Number"] if len(inv["Number"])!=0]
    project_name_last_inv_list.append({
        "Project Name": data_json["Project Info"]["Project"]["Project Name"],
        "Last Invoice Number": max(inv_list),
        "Include Services": ", ".join(
                    [service["Service"] for service in data_json["Project Info"]["Project"]["Service Type"].values() if
                     service["Include"]]),
        "Invoice List": ", ".join(inv_list)
    })
    return max(inv_list)

# type1_json = convert_asana_name_to_project_number(["P:\\308230-Function Centre, 4A Kenthurst Road, Dural"])

# print(type1_json)
assign_service_map = {
    "Mechanical": "Mechanical Service",
    "Review": "Mechanical Review",
    "Hydraulic": "Hydraulic Service",
    "Fire": "Fire Service",
    "Electrical": "Electrical Service",
    "CFD": "CFD Service",
    "Installation": "Installation",
    "Miscellaneous": "Miscellaneous",
    "Variation": "Variation"
}


pce_inv_2 = "S:\\20240226-Bridge manul match Fee Details\\PCE INV-Manual rectify_Nith.xlsx"
pce_inv_3 = "S:\\20240226-Bridge manul match Fee Details\\PCE INV-Manual rectify_Alex.xlsx"
# excel = win32client.Dispatch("Excel.Application")
# work_book = excel.Workbooks.Open(pce_inv)
# work_sheets = work_book.Worksheets[0]


wb_2 = load_workbook(pce_inv_2)
wb_3 = load_workbook(pce_inv_3)
# print(wb.sheetnames)
# example_sheet = wb['312408']

database_dir = conf["database_dir"]
skip_project = ["205000ZK", "301000AJ", "306000AB", "307000AC", "308000AZ", "309000AG", "309000AJ", "310000ZM",
                "311000ZG", "311000ZP", "311000ZR", "312000AT", "312000AV", "312000AX", "312000BE", "312000BF",
                "312000BJ"]

last_inv_quotation_map = generate_last_inv_to_quotation_map(database_dir)


# skip_invoice = ["206119", "206120", "206121", "206123", "304098", "312399"]
processed_project = []
count=0
all_sheetnames = wb_2.sheetnames




for sheetname in all_sheetnames:
    if sheetname.startswith("2"):
        wb = wb_2
    elif sheetname.startswith("3"):
        wb = wb_3
    else:
        break
    # if sheetname in skip_invoice:
    #     continue
    # elif sheetname[0:3] > "206":
    #     break
    print(f"Processing Sheet {sheetname}")
    assert sheetname in last_inv_quotation_map.keys()
    quotation = last_inv_quotation_map[sheetname]
    if quotation in skip_project:
        print(f"Skipping {sheetname} since quotation {quotation} already manually match")
        continue
    processed_project.append(quotation)
    print(f"Invoice {sheetname} with Quotation {quotation}")
    data_dir = os.path.join(database_dir, quotation, "data.json")
    data_json = json.load(open(data_dir))

    inv_excel = wb[sheetname]
    current_row = 11
    excel_total_amount = 0
    data_total_amount = float(data_json["Invoices"]["Fee"])
    for service in conf["invoice_list"]:
        for content in data_json["Invoices"]["Details"][service]["Content"]:
            content["Fee"] = ""
            content["Service"] = ""
            content["in.GST"] = ""
            content["Number"] = "None"
    while True:
        item = inv_excel[f"A{current_row}"].value
        if current_row == 38:
            break
        if not item is None:
            amount = inv_excel[f"G{current_row}"].value
            if not amount is None and (str(amount).startswith("=") or str(amount).startswith("Paid")):
                print(sheetname)
            if isfloat(amount):
                excel_total_amount += float(amount)
                first_word = item.split(" ")[0].split("-")[0]
                # if first_word in assign_service_map.keys():
                service = assign_service_map[first_word]
                # assert service in [service["Service"] for service in
                #                    data_json["Project Info"]["Project"]["Service Type"].values() if service["Include"]]
                index = None
                for i in range(4):
                    if len(data_json["Invoices"]["Details"][service]["Content"][i]["Fee"]) == 0:
                        index = i
                        break
                if index is None:
                    raise Exception(f"Invoice {quotation} with Over 4 {service} when adding itme {item}")
                data_json["Invoices"]["Details"][service]["Content"][index]["Service"] = item
                data_json["Invoices"]["Details"][service]["Content"][index]["Fee"] = str(amount)
                data_json["Invoices"]["Details"][service]["Content"][index]["in.GST"] = str(round(amount * 1.1, 2))

        current_row += 1

    check_valid_project(data_json)
    # raise Exception(f"Invoice {last_inv_number} with different amounts, the amount should be {data_total_amount}, the excel sum is {excel_total_amount}")
    with open(data_dir, "w") as f:
        json_object = json.dumps(data_json, indent=4)
        f.write(json_object)
# report = []

# def return_scope(data_json):
#     res = []
#     for service in conf["invoice_list"]:
#         for content in data_json["Invoices"]["Details"][service]["Content"]:
#             if len(content["Service"]) != 0 or len(content["Fee"]) != 0:
#                 res.append(f"{content['Service']}\t{content['Fee']}")
#     return res


filepath = "S:\\20240226-Bridge manul match Fee Details\\report.xlsx"
report_wb = openpyxl.Workbook()
report_ws = report_wb.active
report_ws.title = "report"

report_ws["A1"] = "Quotation"
report_ws["B1"] = "Project Number"
report_ws["C1"] = "All Invoices"
report_ws["D1"] = "First Invoice"
report_ws["E1"] = "Last Invoice"
report_ws["F1"] = "Project Address"
report_ws["G1"] = "Project Service/Scope"
report_ws["H1"] = "Project Total/Scope Fee"
report_ws["I1"] = "Total From the Last Invoice"
report_ws["J1"] = "Scope match"


def check_fee(data_json):
    for key, service in data_json["Invoices"]["Details"].items():
        if key == "Variation" or not service["Include"]:
            continue
        if service["Fee"] == "0" or len(service["Fee"]) == 0:
            return False
    return True
current_row = 2
for quotation in processed_project:
    initial_row = current_row
    data_json = json.load(open(os.path.join(database_dir, quotation, "data.json")))
    report_ws[f"A{current_row}"] = quotation
    report_ws[f"B{current_row}"] = data_json["Project Info"]["Project"]["Project Number"]
    all_invoices = [inv["Number"] for inv in data_json["Invoices Number"] if len(inv["Number"]) !=0]
    report_ws[f"C{current_row}"] = ", ".join(all_invoices)
    report_ws[f"D{current_row}"] = "" if len(all_invoices) == 0 else data_json["Invoices Number"][0]["Number"]
    report_ws[f"E{current_row}"] = get_last_inv_number(data_json)

    if data_json["Project Info"]["Project"]["Project Number"] == "":
        project_address = "P:\\"+data_json["Project Info"]["Project"]["Quotation Number"]+"-"+data_json["Project Info"]["Project"]["Project Name"]
    else:
        project_address = "P:\\"+data_json["Project Info"]["Project"]["Project Number"]+"-"+data_json["Project Info"]["Project"]["Project Name"]
    report_ws[f"F{current_row}"] = project_address
    report_ws[f"G{current_row}"] = ", ".join([service["Service"] for service in data_json["Project Info"]["Project"]["Service Type"].values() if service["Include"]])
    report_ws[f"H{current_row}"] = float(data_json["Invoices"]["Fee"])
    current_row += 1
    sub_total = 0
    match = check_fee(data_json)
    for service in conf["invoice_list"]:
        for content in data_json["Invoices"]["Details"][service]["Content"]:
            if len(content["Service"]) != 0 or len(content["Fee"]) != 0:
                report_ws[f"G{current_row}"] = content['Service']
                report_ws[f"H{current_row}"] = float(content['Fee'])
                report_ws[f"J{current_row}"] = match
                sub_total += float(content["Fee"])
                current_row += 1
    report_ws[f"I{initial_row}"] = float(sub_total)
    report_ws[f"J{initial_row}"] = match
report_wb.save(filepath)
    # report.append({
    #         "Quotation": quotation,
    #         "Last Invoice": get_last_inv_number(data_json),
    #         "Project Number": data_json["Project Info"]["Project"]["Project Number"],
    #         "Project Name": data_json["Project Info"]["Project"]["Project Name"],
    #         "Project Include Services": ", ".join([service["Service"] for service in data_json["Project Info"]["Project"]["Service Type"].values() if service["Include"]]),
    #         "Project Total": data_json["Invoices"]["Fee"],
    #         "Scope": return_scope(data_json)
    #     }
    # )
# with open(os.path.join(database_dir, "report.json"), "w") as f:
#     json_object = json.dumps(report, indent=4)
#     f.write(json_object)


# for dir in os.listdir(database_dir):
#     if os.path.isdir(os.path.join(database_dir, dir)):
#         # if str(dir).startswith("2") or str(dir).startswith("3"):
#         if dir in ["201000AH", "201000AE", "201000AI", "201000AJ"]:
#             if str(dir) in skip_project:
#                 continue
#             print(dir)
#             data_dir = os.path.join(database_dir, dir, "data.json")
#             data_json = json.load(open(data_dir))
#             last_inv_number = get_last_inv_number(data_json)
#             inv_excel = wb[last_inv_number]
#             current_row = 11
#             excel_total_amount = 0
#             data_total_amount = float(data_json["Invoices"]["Fee"])
#             for service in conf["invoice_list"]:
#                 for content in data_json["Invoices"]["Details"][service]["Content"]:
#                     content["Fee"] = ""
#                     content["Service"] = ""
#                     content["in.GST"] = ""
#                     content["Number"] = "None"
#             while True:
#                 item = inv_excel[f"A{current_row}"].value
#                 if current_row == 38:
#                     break
#                 if not item is None:
#                     amount = inv_excel[f"G{current_row}"].value
#                     if isfloat(amount):
#                         excel_total_amount += float(amount)
#                         first_word = item.split(" ")[0].split("-")[0]
#                         # if first_word in assign_service_map.keys():
#                         service = assign_service_map[first_word]
#                         assert service in [service["Service"] for service in data_json["Project Info"]["Project"]["Service Type"].values() if service["Include"]]
#                         # else:
#                         #     service = "Variation"
#                         index = None
#
#                         for i in range(4):
#                             if len(data_json["Invoices"]["Details"][service]["Content"][i]["Fee"]) == 0:
#                                 index = i
#                                 break
#                         data_json["Invoices"]["Details"][service]["Content"][index]["Service"] = item
#                         data_json["Invoices"]["Details"][service]["Content"][index]["Fee"] = str(amount)
#                         data_json["Invoices"]["Details"][service]["Content"][index]["in.GST"] = str(round(amount*1.1, 2))
#                         if index is None:
#                             raise Exception(f"Invoice {last_inv_number} with Over 4 {service} when adding itme {item}")
#                 current_row += 1
#
#
#             check_valid_project(data_json)
#             # raise Exception(f"Invoice {last_inv_number} with different amounts, the amount should be {data_total_amount}, the excel sum is {excel_total_amount}")
#             with open(data_dir, "w") as f:
#                 json_object = json.dumps(data_json, indent=4)
#                 f.write(json_object)
# with open(data_dir, "w") as f:
#     json_object = json.dumps(data_json, indent=4)
#     f.write(json_object)
print("Syncing Complete")