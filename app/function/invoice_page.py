import tkinter as tk
from tkinter import ttk
from function.utility import *
from openpyxl import load_workbook
import os
class InvoicePage(tk.Frame):
    def __init__(self, parent, app):
        tk.Frame.__init__(self, parent)
        self.update_invoice_number = None
        self.parent = parent
        self.app = app
        self.main_frame = tk.Frame(self, bg="blue")
        self.main_frame.pack(fill=tk.BOTH, expand=1, side=tk.LEFT, anchor="nw")

        self.main_canvas = tk.Canvas(self.main_frame)
        self.main_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=1, anchor="nw")

        self.scrollbar = ttk.Scrollbar(self.main_frame, orient=tk.VERTICAL, command=self.main_canvas.yview)
        self.scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        self.main_canvas.config(yscrollcommand=self.scrollbar.set)
        self.main_canvas.bind("<Configure>", lambda e:self.main_canvas.configure(scrollregion=self.main_canvas.bbox("all")))
        self.app.bind("<MouseWheel>", self._on_mousewheel, add="+")
        self.main_context_frame = tk.LabelFrame(self.main_canvas, text="Main Context")
        self.main_canvas.create_window((0, 0), window=self.main_context_frame, anchor="nw")
        self.main_context_frame.bind("<Configure>", self.reset_scrollregion)
        self.items_frame = {}
        self.bills_frame = {}

        self.invoice_main_frame()
        self.bill_main_frame()
        self.profit_main_frame()

    def invoice_main_frame(self):
        self.fee_frame = tk.LabelFrame(self.main_context_frame, text="Invoice Details", font=self.app.font)
        self.fee_frame.pack(side=tk.LEFT,fill=tk.BOTH, expand=1)

        self.invoice_frame = tk.LabelFrame(self.fee_frame)
        tk.Label(self.invoice_frame, width=35, text="Items", font=self.app.font).grid(row=0, column=0)
        tk.Label(self.invoice_frame, width=10, text="ex.GST", font=self.app.font).grid(row=0, column=1)
        tk.Label(self.invoice_frame, width=10, text="in.GST", font=self.app.font).grid(row=1, column=1)
        #for demonstration purpose
        tk.Button(self.invoice_frame,width=10, text="Generate Invoice", command= self.generate_invoice_number, bg="brown", fg="white",
                  font=self.app.font).grid(row=1, column=2)
        tk.Button(self.invoice_frame, width=10, text="Update Xero", command=lambda: update_xero(self.app.data, self.update_invoice_number), bg="brown",
                  fg="white",
                  font=self.app.font).grid(row=1, column=3)
        #
        for i in range(6):
            tk.Label(self.invoice_frame, width=10, text="INV"+str(i+1), font=self.app.font).grid(row=0, column=i+2, sticky="ew")
        self.invoice_frame.grid(row=0, column=0)

        self.invoice_number_frame = tk.LabelFrame(self.fee_frame)
        tk.Label(self.invoice_number_frame, width=10, text="", font=self.app.font).grid(row=0, column=1)
        tk.Label(self.invoice_number_frame, width=35, text="Invoice Number", font=self.app.font).grid(row=0, column=0)
        self.invoice_number_entry = []
        for i in range(6):
            self.invoice_number_entry.append(tk.Entry(self.invoice_number_frame, width=11, font=self.app.font, fg="blue"))
            self.invoice_number_entry[i].grid(row=0, column=i + 2, padx=(0,5), sticky="ew")
        self.invoice_number_frame.grid(row=1, column=0, sticky="ew")

        self.fee_dic = dict()
        self.fee_frame_number = 2

        self.total_frame = tk.LabelFrame(self.fee_frame)
        tk.Label(self.total_frame, width=35, text="Total", font=self.app.font).grid(row=0, column=0)
        self.total_ex_gst_label = tk.Label(self.total_frame, width=10, textvariable=self.app.data["Fee Proposal Page"]["Fee"], font=self.app.font)
        self.total_ex_gst_label.grid(row=0, column=1)
        self.total_in_gst_label = tk.Label(self.total_frame, width=10, textvariable=self.app.data["Fee Proposal Page"]["in.GST"], font=self.app.font)
        self.total_in_gst_label.grid(row=1, column=1)
        self.invoice_list = [tk.StringVar(name="Invoice "+str(i+1), value="0") for i in range(6)]
        self.inGST_invoice_list = [tk.StringVar(name="ingst Invoice "+str(i+1), value="0") for i in range(6)]
        for j in range(6):
            tk.Label(self.total_frame, width=10, textvariable=self.invoice_list[j], font=self.app.font).grid(row=0, column=2+j)
            tk.Label(self.total_frame, width=10, textvariable=self.inGST_invoice_list[j], font=self.app.font).grid(row=1, column=2+j)
        self.total_frame.grid(row=999, column=0, sticky="ew")
        #garbage collection
        self.invoice_list[0].trace("w", lambda a, b, c: self.app._ist_update(self.invoice_list[0],
                                                                             self.inGST_invoice_list[0]))
        self.invoice_list[1].trace("w", lambda a, b, c: self.app._ist_update(self.invoice_list[1],
                                                                             self.inGST_invoice_list[1]))
        self.invoice_list[2].trace("w", lambda a, b, c: self.app._ist_update(self.invoice_list[2],
                                                                             self.inGST_invoice_list[2]))
        self.invoice_list[3].trace("w", lambda a, b, c: self.app._ist_update(self.invoice_list[3],
                                                                             self.inGST_invoice_list[3]))
        self.invoice_list[4].trace("w", lambda a, b, c: self.app._ist_update(self.invoice_list[4],
                                                                             self.inGST_invoice_list[4]))
        self.invoice_list[5].trace("w", lambda a, b, c: self.app._ist_update(self.invoice_list[5],
                                                                             self.inGST_invoice_list[5]))

    def bill_main_frame(self):
        self.bill_frame = tk.LabelFrame(self.main_context_frame, text="Bill Details", font=self.app.font)
        self.bill_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=1)

        self.invoice_frame = tk.LabelFrame(self.bill_frame)
        tk.Label(self.invoice_frame, width=35, text="Bills", font=self.app.font).grid(row=0, column=0)
        tk.Label(self.invoice_frame, width=10, text="ex.GST", font=self.app.font).grid(row=0, column=1)
        tk.Label(self.invoice_frame, width=10, text="in.GST", font=self.app.font).grid(row=1, column=1)
        for i in range(6):
            tk.Label(self.invoice_frame, width=10, text="BILL" + str(i + 1), font=self.app.font).grid(row=0,
                                                                                                     column=i + 2)
        self.invoice_frame.grid(row=0, column=0, sticky="ew")

        self.invoice_number_frame = tk.LabelFrame(self.bill_frame)
        tk.Label(self.invoice_number_frame, width=10, text="", font=self.app.font).grid(row=0, column=1)
        tk.Label(self.invoice_number_frame, width=35, text="Bill Number", font=self.app.font).grid(row=0, column=0)
        for i in range(6):
            tk.Entry(self.invoice_number_frame, width=11, font=self.app.font, fg="blue").grid(row=0, column=i + 2,
                                                                                              padx=(0, 5), sticky="ew")
        self.invoice_number_frame.grid(row=1, column=0, sticky="ew")

        self.bill_data_dic = dict()
        self.bill_fee_total = tk.StringVar(name="bill fee total")
        self.bill_fee_isgst_total = tk.StringVar(name="bill ingst fee total")
        self.bill_dic = dict()
        self.bill_frame_number = 2

        self.total_frame = tk.LabelFrame(self.bill_frame)
        tk.Label(self.total_frame, width=35, text=" Bill Total", font=self.app.font).grid(row=0, column=0)
        self.total_ex_gst_label = tk.Label(self.total_frame, width=10,
                                           textvariable=self.bill_fee_total, font=self.app.font)
        self.total_ex_gst_label.grid(row=0, column=1)
        self.total_in_gst_label = tk.Label(self.total_frame, width=10,
                                           textvariable=self.bill_fee_isgst_total,
                                           font=self.app.font)
        self.total_in_gst_label.grid(row=1, column=1)
        self.bill_list = [tk.StringVar(name="Bill " + str(i + 1), value="0") for i in range(6)]
        self.inGST_bill_list = [tk.StringVar(name="ingst Bill " + str(i + 1), value="0") for i in range(6)]
        for j in range(6):
            tk.Label(self.total_frame, width=10, textvariable=self.bill_list[j], font=self.app.font).grid(row=0,
                                                                                                             column=2 + j)
            tk.Label(self.total_frame, width=10, textvariable=self.inGST_bill_list[j], font=self.app.font).grid(
                row=1, column=2 + j)
        self.total_frame.grid(row=999, column=0, sticky="ew")
        self.bill_list[0].trace("w", lambda a, b, c: self.app._ist_update(self.bill_list[0],
                                                                             self.inGST_bill_list[0]))
        self.bill_list[1].trace("w", lambda a, b, c: self.app._ist_update(self.bill_list[1],
                                                                             self.inGST_bill_list[1]))
        self.bill_list[2].trace("w", lambda a, b, c: self.app._ist_update(self.bill_list[2],
                                                                             self.inGST_bill_list[2]))
        self.bill_list[3].trace("w", lambda a, b, c: self.app._ist_update(self.bill_list[3],
                                                                             self.inGST_bill_list[3]))
        self.bill_list[4].trace("w", lambda a, b, c: self.app._ist_update(self.bill_list[4],
                                                                             self.inGST_bill_list[4]))
        self.bill_list[5].trace("w", lambda a, b, c: self.app._ist_update(self.bill_list[5],
                                                                             self.inGST_bill_list[5]))
    def profit_main_frame(self):
        self.profit_frame = tk.LabelFrame(self.main_context_frame, text="Profit Details", font=self.app.font)
        self.profit_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=1)

        self.profit_title_frame = tk.LabelFrame(self.profit_frame)
        tk.Label(self.profit_title_frame, width=10, text="ex.GST", font=self.app.font).grid(row=0, column=0)
        tk.Label(self.profit_title_frame, width=10, text="in.GST", font=self.app.font).grid(row=1, column=0)
        self.profit_title_frame.grid(row=0, column=0, sticky="ew")

        _ = tk.LabelFrame(self.profit_frame)
        tk.Label(_, text="").grid(row=0, column=0, pady=(0, 3))
        _.grid(row=1, column=0, sticky="ew")
        
        self.profit_data_dic = dict()
        self.profit_dic = dict()
        self.profits_frame = dict()
        self.profit_frame_number = 2

        self.total_profit_frame = tk.LabelFrame(self.profit_frame)
        self.total_profit_var = tk.StringVar(name="Total Profit", value = "")
        self.total_in_profit_var = tk.StringVar(name="Total ingst Profit", value="")
        tk.Label(self.total_profit_frame, width=10, textvariable=self.total_profit_var, font=self.app.font).grid(row=0, column=0)
        tk.Label(self.total_profit_frame, width=10, textvariable=self.total_in_profit_var, font=self.app.font).grid(row=1, column=0)

        self.total_profit_frame.grid(row=999, column=0, sticky="ew")
        

    def update_fee(self, var):
        if var.get() == True:
            if not var._name in self.fee_dic.keys():
                self.items_frame[var._name] = tk.LabelFrame(self.fee_frame)
                self.fee_dic[var._name] = {
                    "Service": tk.Label(self.items_frame[var._name],
                                        text=var._name + " design and documentation",
                                        width=35,
                                        font=self.app.font),
                    "Fee": tk.Label(self.items_frame[var._name],
                                    textvariable=self.app.data["Fee Proposal Page"]["Details"][var._name]["Fee"],
                                    width=10,
                                    font=self.app.font),
                    "in.GST": tk.Label(self.items_frame[var._name],
                                       textvariable=self.app.data["Fee Proposal Page"]["Details"][var._name]["in.GST"],
                                       width=10,
                                       font=self.app.font),
                    "Position": 0,
                    "Invoice":[tk.Radiobutton(self.items_frame[var._name],
                                              width=8,
                                              variable=self.app.data["Fee Proposal Page"]["Details"][var._name]["Invoice"],
                                              value="INV"+str(i+1)) for i in range(6)],
                    "expand frames": [tk.LabelFrame(self.fee_frame) for _ in range(4)]
                }
                self.fee_dic[var._name]["expand frame"]=[
                    {
                        "Service":tk.Label(self.fee_dic[var._name]["expand frames"][i],
                                           width=35,
                                           textvariable=self.app.data["Fee Proposal Page"]["Details"][var._name]["Context"][i]["Service"],
                                           font=self.app.font),
                        "Fee":tk.Label(self.fee_dic[var._name]["expand frames"][i],
                                       width=10,
                                       textvariable=self.app.data["Fee Proposal Page"]["Details"][var._name]["Context"][i]["Fee"],
                                       font=self.app.font),
                        "in.GST": tk.Label(self.fee_dic[var._name]["expand frames"][i],
                                           width=10,
                                           textvariable=self.app.data["Fee Proposal Page"]["Details"][var._name]["Context"][i]["in.GST"],
                                           font=self.app.font),
                        "Invoice": [tk.Radiobutton(self.fee_dic[var._name]["expand frames"][i],
                                                   width=8,
                                                   variable=self.app.data["Fee Proposal Page"]["Details"][var._name]["Context"][i]["Invoice"],
                                                   value="INV" + str(j + 1)) for j in range(6)]
                    }for i in range(3)
                ]
                self.fee_dic[var._name]["Total"]=tk.Label(self.fee_dic[var._name]["expand frames"][3], width=35, text=var._name + " Total",
                                                          font=self.app.font)
                self.fee_dic[var._name]["Total Fee"]=tk.Label(self.fee_dic[var._name]["expand frames"][3], width=10,
                                                              textvariable=self.app.data["Fee Proposal Page"]["Details"][var._name]["Fee"],
                                                              font=self.app.font)
                self.fee_dic[var._name]["Total in.GST"]=tk.Label(self.fee_dic[var._name]["expand frames"][3], width=10,
                                                                 textvariable=self.app.data["Fee Proposal Page"]["Details"][var._name][
                                                                     "in.GST"],
                                                                 font=self.app.font)

                self.app.data["Fee Proposal Page"]["Details"][var._name]["Expanded"].trace("w",lambda a, b, c: self.expand(var._name))
                self.app.data["Fee Proposal Page"]["Details"][var._name]["Invoice"].trace("w", lambda a,b,c:self.app.update_invoice_sum(self.app.data["Fee Proposal Page"]["Details"], self.invoice_list))
                self.app.data["Fee Proposal Page"]["Details"][var._name]["Fee"].trace("w", lambda a,b,c:self.app.update_invoice_sum(self.app.data["Fee Proposal Page"]["Details"], self.invoice_list))

                #garbage collect
                self.app.data["Fee Proposal Page"]["Details"][var._name]["Context"][0]["Service"].trace("w", lambda a,b,c: self.update_radiobutton(self.app.data["Fee Proposal Page"]["Details"][var._name]["Context"][0]["Service"]))
                self.app.data["Fee Proposal Page"]["Details"][var._name]["Context"][1]["Service"].trace("w", lambda a,b,c: self.update_radiobutton(self.app.data["Fee Proposal Page"]["Details"][var._name]["Context"][1]["Service"]))
                self.app.data["Fee Proposal Page"]["Details"][var._name]["Context"][2]["Service"].trace("w", lambda a,b,c: self.update_radiobutton(self.app.data["Fee Proposal Page"]["Details"][var._name]["Context"][2]["Service"]))

                for i in range(3):
                    self.app.data["Fee Proposal Page"]["Details"][var._name]["Context"][i]["Invoice"].trace("w", lambda a, b,
                                                                                                          c: self.app.update_invoice_sum(
                        self.app.data["Fee Proposal Page"]["Details"], self.invoice_list))


                for i in range(3):
                    self.fee_dic[var._name]["expand frame"][i]["Service"].grid(row=0, column=0)
                    self.fee_dic[var._name]["expand frame"][i]["Fee"].grid(row=0, column=1)
                    self.fee_dic[var._name]["expand frame"][i]["in.GST"].grid(row=1, column=1, padx=(0, 2))
                self.fee_dic[var._name]["Total"].grid(row=0, column=0)
                self.fee_dic[var._name]["Total Fee"].grid(row=0, column=1)
                self.fee_dic[var._name]["Total in.GST"].grid(row=1, column=1)

                self.fee_dic[var._name]["Service"].grid(row=0, column=0)
                self.fee_dic[var._name]["Fee"].grid(row=0, column=1)
                self.fee_dic[var._name]["in.GST"].grid(row=1, column=1, padx=(0, 2))

                for i in range(6):
                    self.fee_dic[var._name]["Invoice"][i].grid(row=0, column=3+i, padx=(2, 0), rowspan=2)




            self.fee_dic[var._name]["Position"] = str(self.fee_frame_number)
            self.items_frame[var._name].grid(row=self.fee_frame_number, column=0, sticky="ew")
            self.fee_frame_number += 5
        else:
            self.app.data["Fee Proposal Page"]["Details"][var._name]["Fee"].set("")
            if self.app.data["Fee Proposal Page"]["Details"][var._name]["Expanded"].get():
                self.app.data["Fee Proposal Page"]["Details"][var._name]["Expanded"].set(False)
            self.items_frame[var._name].grid_forget()
    
    def update_bill(self, var):
        if var.get() == True:
            if not var._name in self.bill_dic.keys():
                
                self.bills_frame[var._name] = tk.LabelFrame(self.bill_frame)
                self.bill_data_dic[var._name] = {
                    "on":tk.BooleanVar(name=var._name+" bill on", value=True),
                    "Expanded":tk.BooleanVar(name=var._name + " bill Expand", value=False),
                    "Fee":tk.StringVar(name=var._name + " bill Fee"),
                    "in.GST":tk.StringVar(name=var._name + " bill in.GST"),
                    "Invoice":tk.StringVar(name=var._name + " Bill", value="None"),
                    "Context":[
                        {
                            "Service":tk.StringVar(name=var._name+" bill "+str(i)),
                            "Fee":tk.StringVar(name=var._name+" bill Fee "+str(i)),
                            "in.GST":tk.StringVar(name=var._name+" bill in.GST "+str(i)),
                            "Invoice": tk.StringVar(name=var._name +str(i)+ " Bill", value="None")
                        } for i in range(3)
                    ]
                }

                self.bill_dic[var._name] = {
                    "Service": tk.Entry(self.bills_frame[var._name],
                                        width=35,
                                        font=self.app.font,
                                        fg="blue"),
                    "Fee": tk.Entry(self.bills_frame[var._name],
                                    textvariable=self.bill_data_dic[var._name]["Fee"],
                                    width=10,
                                    font=self.app.font,
                                    fg="blue"),
                    "in.GST": tk.Label(self.bills_frame[var._name],
                                       textvariable=self.bill_data_dic[var._name]["in.GST"],
                                       width=10,
                                       font=self.app.font),
                    "Position": 0,
                    "Invoice":[tk.Radiobutton(self.bills_frame[var._name],
                                              width=8,
                                              variable=self.bill_data_dic[var._name]["Invoice"],
                                              value="INV"+str(i+1)) for i in range(6)],
                    "expand frames": [tk.LabelFrame(self.bill_frame) for _ in range(4)]
                }
                self.bill_dic[var._name]["expand frame"]=[
                    {
                        "Service":tk.Entry(self.bill_dic[var._name]["expand frames"][i],
                                           width=35,
                                           textvariable=self.bill_data_dic[var._name]["Context"][i]["Service"],
                                           font=self.app.font,
                                           fg="blue"),
                        "Fee":tk.Entry(self.bill_dic[var._name]["expand frames"][i],
                                       width=10,
                                       textvariable=self.bill_data_dic[var._name]["Context"][i]["Fee"],
                                       font=self.app.font,
                                       fg="blue"),
                        "in.GST": tk.Label(self.bill_dic[var._name]["expand frames"][i],
                                           width=10,
                                           textvariable=self.bill_data_dic[var._name]["Context"][i]["in.GST"],
                                           font=self.app.font),
                        "Invoice": [tk.Radiobutton(self.bill_dic[var._name]["expand frames"][i],
                                                   width=8,
                                                   variable=self.bill_data_dic[var._name]["Context"][i]["Invoice"],
                                                   value="INV" + str(j + 1)) for j in range(6)]
                    }for i in range(3)
                ]

                self.bill_dic[var._name]["Total"]=tk.Label(self.bill_dic[var._name]["expand frames"][3], width=35, text=var._name+" Bill Total",
                                                          font=self.app.font)
                self.bill_dic[var._name]["Total Fee"]=tk.Label(self.bill_dic[var._name]["expand frames"][3], width=10,
                                                              textvariable=self.bill_data_dic[var._name]["Fee"],
                                                              font=self.app.font)
                self.bill_dic[var._name]["Total in.GST"]=tk.Label(self.bill_dic[var._name]["expand frames"][3], width=10,
                                                                 textvariable=self.bill_data_dic[var._name][
                                                                     "in.GST"],
                                                                 font=self.app.font)

                #trace function start

                self.app.data["Fee Proposal Page"]["Details"][var._name]["Expanded"].trace("w",lambda a, b, c: self.bill_expand(var._name))
                self.bill_data_dic[var._name]["Fee"].trace("w", lambda a,b,c:self.app._ist_update(
                    self.bill_data_dic[var._name]["Fee"], self.bill_data_dic[var._name]["in.GST"]))

                self.bill_data_dic[var._name]["Fee"].trace("w", lambda a,b,c: self.app._sum_update(
                    [value["Fee"] for value in self.bill_data_dic.values()], self.bill_fee_total))

                self.bill_data_dic[var._name]["in.GST"].trace("w", lambda a,b,c: self.app._sum_update(
                    [value["in.GST"] for value in self.bill_data_dic.values()], self.bill_fee_isgst_total))



                #garbar collect need to fix
                ist_fun_0=lambda a, b, c: self.app._ist_update(
                    self.bill_data_dic[var._name]["Context"][0]["Fee"],
                    self.bill_data_dic[var._name]["Context"][0]["in.GST"])
                self.bill_data_dic[var._name]["Context"][0]["Fee"].trace("w", ist_fun_0)
                ist_fun_1=lambda a, b, c: self.app._ist_update(
                    self.bill_data_dic[var._name]["Context"][1]["Fee"],
                    self.bill_data_dic[var._name]["Context"][1]["in.GST"])
                self.bill_data_dic[var._name]["Context"][1]["Fee"].trace("w", ist_fun_1)
                ist_fun_2=lambda a, b, c: self.app._ist_update(
                    self.bill_data_dic[var._name]["Context"][2]["Fee"],
                    self.bill_data_dic[var._name]["Context"][2]["in.GST"])
                self.bill_data_dic[var._name]["Context"][2]["Fee"].trace("w", ist_fun_2)
                for i in range(3):
                    sum_fun = lambda a, b, c: self.app._sum_update([item["Fee"] for item in self.bill_data_dic[var._name]["Context"]],
                                                                   self.bill_data_dic[var._name]["Fee"])
                    self.bill_data_dic[var._name]["Context"][i]["Fee"].trace("w", sum_fun)


                self.bill_data_dic[var._name]["Invoice"].trace("w", lambda a,b,c:self.app.update_invoice_sum(self.bill_data_dic, self.bill_list))
                self.bill_data_dic[var._name]["Fee"].trace("w", lambda a,b,c:self.app.update_invoice_sum(self.bill_data_dic, self.bill_list))

                #garbage collect
                self.bill_data_dic[var._name]["Context"][0]["Service"].trace("w", lambda a,b,c: self.update_bill_radiobutton(self.bill_data_dic[var._name]["Context"][0]["Service"]))
                self.bill_data_dic[var._name]["Context"][1]["Service"].trace("w", lambda a,b,c: self.update_bill_radiobutton(self.bill_data_dic[var._name]["Context"][1]["Service"]))
                self.bill_data_dic[var._name]["Context"][2]["Service"].trace("w", lambda a,b,c: self.update_bill_radiobutton(self.bill_data_dic[var._name]["Context"][2]["Service"]))

                for i in range(3):
                    self.bill_data_dic[var._name]["Context"][i]["Invoice"].trace("w", lambda a, b,c: self.app.update_invoice_sum(self.bill_data_dic, self.bill_list))


                for i in range(3):
                    self.bill_dic[var._name]["expand frame"][i]["Service"].grid(row=0, column=0, padx=20, pady=(2, 0))
                    self.bill_dic[var._name]["expand frame"][i]["Fee"].grid(row=0, column=1)
                    self.bill_dic[var._name]["expand frame"][i]["in.GST"].grid(row=1, column=1, padx=(0, 2))
                self.bill_dic[var._name]["Total"].grid(row=0, column=0)
                self.bill_dic[var._name]["Total Fee"].grid(row=0, column=1)
                self.bill_dic[var._name]["Total in.GST"].grid(row=1, column=1)

                self.bill_dic[var._name]["Service"].grid(row=0, column=0, padx=20, pady=(2, 0))
                self.bill_dic[var._name]["Fee"].grid(row=0, column=1)
                self.bill_dic[var._name]["in.GST"].grid(row=1, column=1, padx=(0, 2))

                for i in range(6):
                    self.bill_dic[var._name]["Invoice"][i].grid(row=0, column=3+i, padx=(2, 0), rowspan=2)




            self.bill_dic[var._name]["Position"] = str(self.bill_frame_number)
            self.bills_frame[var._name].grid(row=self.bill_frame_number, column=0, sticky="ew")
            self.bill_frame_number += 5
        else:
            self.bill_data_dic[var._name]["Fee"].set("")
            if self.bill_data_dic[var._name]["Expanded"].get():
                self.bill_data_dic[var._name]["Expanded"].set(False)
            self.bills_frame[var._name].grid_forget()
    
    def update_profit(self, var):
        if var.get() == True:
            if not var._name in self.profit_dic.keys():
                self.profits_frame[var._name] = tk.LabelFrame(self.profit_frame)
                
                self.profit_data_dic[var._name] = {
                    "on":tk.BooleanVar(name=var._name+" profit on", value=True),
                    "Expanded":tk.BooleanVar(name=var._name + " profit Expand", value=False),
                    "Fee":tk.StringVar(name=var._name + " profit Fee"),
                    "in.GST":tk.StringVar(name=var._name + " profit in.GST"),
                    "Context":[
                        {
                            "Fee":tk.StringVar(name=var._name+" profit Fee "+str(i)),
                            "in.GST":tk.StringVar(name=var._name+" profit in.GST "+str(i)),
                        } for i in range(3)
                    ]
                }
                
                self.profit_dic[var._name] = {
                    "Fee": tk.Label(self.profits_frame[var._name],
                                    textvariable=self.profit_data_dic[var._name]["Fee"],
                                    width=10,
                                    font=self.app.font),
                    "in.GST": tk.Label(self.profits_frame[var._name],
                                       textvariable=self.profit_data_dic[var._name]["in.GST"],
                                       width=10,
                                       font=self.app.font),
                    "Position": 0,
                    "expand frames": [tk.LabelFrame(self.profit_frame) for _ in range(4)]
                }
                self.profit_dic[var._name]["expand frame"] = [
                    {
                        "Fee": tk.Label(self.profit_dic[var._name]["expand frames"][i],
                                        width=10,
                                        textvariable=self.profit_data_dic[var._name]["Context"][i]["Fee"],
                                        font=self.app.font),
                        "in.GST": tk.Label(self.profit_dic[var._name]["expand frames"][i],
                                           width=10,
                                           textvariable=self.profit_data_dic[var._name]["Context"][i]["in.GST"],
                                           font=self.app.font)
                    } for i in range(3)
                ]
                self.profit_dic[var._name]["Total Fee"]=tk.Label(self.profit_dic[var._name]["expand frames"][3], width=10,
                                                              textvariable=self.profit_data_dic[var._name]["Fee"],
                                                              font=self.app.font)
                self.profit_dic[var._name]["Total in.GST"]=tk.Label(self.profit_dic[var._name]["expand frames"][3], width=10,
                                                                 textvariable=self.profit_data_dic[var._name][
                                                                     "in.GST"],
                                                                 font=self.app.font)
                ###
                #trace function here
                self.app.data["Fee Proposal Page"]["Details"][var._name]["Expanded"].trace("w", lambda a, b, c: self.profit_expand(var._name))
                self.app.data["Fee Proposal Page"]["Details"][var._name]["Fee"].trace("w", self._update_profit)
                self.bill_data_dic[var._name]["Fee"].trace("w", self._update_profit)
                ###
                for i in range(3):
                    self.profit_dic[var._name]["expand frame"][i]["Fee"].grid(row=0, column=0)
                    self.profit_dic[var._name]["expand frame"][i]["in.GST"].grid(row=1, column=0, padx=(0, 2))
                self.profit_dic[var._name]["Total Fee"].grid(row=0, column=0)
                self.profit_dic[var._name]["Total in.GST"].grid(row=1, column=0)

                self.profit_dic[var._name]["Fee"].grid(row=0, column=0)
                self.profit_dic[var._name]["in.GST"].grid(row=1, column=0, padx=(0, 2))

            self.profit_dic[var._name]["Position"] = str(self.profit_frame_number)
            self.profits_frame[var._name].grid(row=self.profit_frame_number, column=0, sticky="ew")
            self.profit_frame_number += 5
        else:
            self.profit_data_dic[var._name]["Fee"].set("")
            if self.profit_data_dic[var._name]["Expanded"].get():
                self.profit_data_dic[var._name]["Expanded"].set(False)
            self.profits_frame[var._name].grid_forget()

    def expand(self, service):
        if self.app.data["Fee Proposal Page"]["Details"][service]["Expanded"].get():
            for i in range(4):
                self.fee_dic[service]["expand frames"][i].grid(row=int(self.fee_dic[service]["Position"])+i+1, column=0, sticky="ew")

            self.app.data["Fee Proposal Page"]["Details"][service]["Invoice"].set("None")
            for j in range(6):
                self.fee_dic[service]["Invoice"][j].grid_forget()
            self.fee_dic[service]["Fee"].grid_forget()
            self.fee_dic[service]["in.GST"].grid_forget()
        else:
            for i in range(4):
                self.fee_dic[service]["expand frames"][i].grid_forget()
            self.fee_dic[service]["Fee"].grid(row=0, column=1)
            self.fee_dic[service]["in.GST"].grid(row=1, column=1, padx=(0, 2))
            for j in range(3):
                self.app.data["Fee Proposal Page"]["Details"][service]["Context"][j]["Invoice"].set("None")
            for k in range(6):
                self.fee_dic[service]["Invoice"][k].grid(row=0, column=k+2, padx=(2, 0), rowspan=2)

    def bill_expand(self, service):
        if self.app.data["Fee Proposal Page"]["Details"][service]["Expanded"].get():
            self.bill_data_dic[service]["Expanded"].set(True)
            for i in range(4):
                self.bill_dic[service]["expand frames"][i].grid(row=int(self.bill_dic[service]["Position"])+i+1, column=0, sticky="ew")
            self.app.data["Fee Proposal Page"]["Details"][service]["Invoice"].set("None")
            for j in range(6):
                self.bill_dic[service]["Invoice"][j].grid_forget()

            self.bill_data_dic[service]["Fee"].set("")
            self.bill_dic[service]["Service"].delete(0, tk.END)
            self.bill_dic[service]["Service"].insert(0, service+" Bills")
            self.bill_dic[service]["Service"].config(state="readonly", fg="black")
            self.bill_dic[service]["Fee"].grid_forget()
            self.bill_dic[service]["in.GST"].grid_forget()
        else:
            self.bill_data_dic[service]["Expanded"].set(False)
            for i in range(4):
                self.bill_dic[service]["expand frames"][i].grid_forget()
            self.bill_dic[service]["Service"].config(state=tk.NORMAL, fg="blue")
            self.bill_dic[service]["Service"].delete(0, tk.END)
            self.bill_dic[service]["Fee"].grid(row=0, column=1)
            self.bill_dic[service]["in.GST"].grid(row=1, column=1, padx=(0, 2))
            for j in range(3):
                self.bill_data_dic[service]["Context"][j]["Service"].set("")
                self.bill_data_dic[service]["Context"][j]["Fee"].set("")
                self.bill_data_dic[service]["Context"][j]["Invoice"].set("None")
            for k in range(6):
                self.bill_dic[service]["Invoice"][k].grid(row=0, column=k+2, padx=(2, 0), rowspan=2)

    def profit_expand(self, service):
        if self.app.data["Fee Proposal Page"]["Details"][service]["Expanded"].get():
            self.profit_data_dic[service]["Expanded"].set(True)
            for i in range(4):
                self.profit_dic[service]["expand frames"][i].grid(row=int(self.profit_dic[service]["Position"]) + i + 1,
                                                                  column=0, sticky="ew")
            self.profit_data_dic[service]["Fee"].set("")
            self.profits_frame[service].config(height=28)
            self.profit_dic[service]["Fee"].grid_forget()
            self.profit_dic[service]["in.GST"].grid_forget()
        else:
            self.profit_data_dic[service]["Expanded"].set(False)
            self.profits_frame[service].config(height=56)
            for i in range(4):
                self.profit_dic[service]["expand frames"][i].grid_forget()
            self.profit_dic[service]["Fee"].grid(row=0, column=1)
            self.profit_dic[service]["in.GST"].grid(row=1, column=1, padx=(0, 2))
            for j in range(3):
                self.profit_data_dic[service]["Context"][j]["Fee"].set("")
    def update_radiobutton(self, var):
        service, index = var._name.split(" item ")
        index = int(index)
        if len(var.get()) == 0:
            for i in range(6):
                self.fee_dic[service]["expand frame"][index]["Invoice"][i].grid_forget()
        else:
            for i in range(6):
                self.fee_dic[service]["expand frame"][index]["Invoice"][i].grid(row=0, column=i + 2, padx=(0, 2),rowspan=2)

    def update_bill_radiobutton(self, var):
        service, index = var._name.split(" bill ")
        index = int(index)
        if len(var.get()) == 0:
            for i in range(6):
                self.bill_dic[service]["expand frame"][index]["Invoice"][i].grid_forget()
        else:
            for i in range(6):
                self.bill_dic[service]["expand frame"][index]["Invoice"][i].grid(row=0, column=i + 2, padx=(0, 2),rowspan=2)


    def _update_profit(self, *args):
        Total = 0
        inGST_Total = 0
        for service_name, service in self.app.data["Fee Proposal Page"]["Details"].items():
            if not service["on"].get():
                continue
            if service["Expanded"].get():
                for i in range(3):
                    try:
                        invoice = service["Context"][i]["Fee"].get() if service["Context"][i]["Fee"].get() != "" else 0
                        bill = self.bill_data_dic[service_name]["Context"][i]["Fee"].get() if self.bill_data_dic[service_name]["Context"][i]["Fee"].get() != "" else 0
                        self.profit_data_dic[service_name]["Context"][i]["Fee"].set(int(invoice) - int(bill))
                        self.profit_data_dic[service_name]["Context"][i]["in.GST"].set(int(float(invoice)*1.1) - int(float(bill)*1.1))
                    except ValueError:
                        self.profit_data_dic[service_name]["Context"][i]["Fee"].set("Error")
                        self.profit_data_dic[service_name]["Context"][i]["in.GST"].set("Error")
            try:
                invoice = service["Fee"].get() if service["Fee"].get() != "" else 0
                bill = self.bill_data_dic[service_name]["Fee"].get() if self.bill_data_dic[service_name]["Fee"].get() != "" else 0
                self.profit_data_dic[service_name]["Fee"].set(int(invoice) - int(bill))
                self.profit_data_dic[service_name]["in.GST"].set(int(float(invoice)*1.1) - int(float(bill)*1.1))
                Total += int(invoice) - int(bill)
                inGST_Total += int(float(invoice)*1.1) - int(float(bill)*1.1)
            except ValueError:
                self.profit_data_dic[service_name]["Fee"].set("Error")
                self.profit_data_dic[service_name]["in.GST"].set("Error")
        self.total_profit_var.set(str(Total))
        self.total_in_profit_var.set(str(inGST_Total))


    def _on_mousewheel(self, event):
        self.main_canvas.yview_scroll(int(-1*(event.delta/120)), "units")

    def reset_scrollregion(self, event):
        self.main_canvas.configure(scrollregion=self.main_canvas.bbox("all"))


    def generate_invoice_number(self):
        wb = load_workbook(os.getcwd()+"\\PCE INV.xlsx")
        cur_number = max([int(num) for num in wb.sheetnames if num.isdigit()])+1
        self.invoice_number_entry[0].delete(0, tk.END)
        self.invoice_number_entry[0].insert(0,str(cur_number))
        self.invoice_number_entry[0].config(state=tk.DISABLED)
        self.update_invoice_number = str(cur_number)
        # sheet = wb.create_sheet(str(cur_number+1), -1)
        # sheet = wb.copy_worksheet(wb[str(cur_number)])
        # sheet = wb[str(cur_number)]
        # print(sheet.title)

